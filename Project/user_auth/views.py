from django.contrib.auth import authenticate, login, logout
from django.http import JsonResponse
from django.shortcuts import redirect, render
from django.urls import reverse
from django.utils.decorators import method_decorator
from django.views import View
from django.views.decorators.csrf import csrf_exempt
from django.views.generic import TemplateView
import pandas as pd
from io import BytesIO
from django.http import HttpResponse
from django.db.models import F
from django.db.models import Q
from rest_framework.parsers import FileUploadParser,MultiPartParser, FormParser
import openpyxl
from .forms import ExcelUploadForm  # Import your form
from django.db import transaction
from django.views.generic import TemplateView
import re
from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_exempt

from rest_framework.permissions import IsAuthenticated
from rest_framework.generics import RetrieveUpdateAPIView
from rest_framework import generics, status
from rest_framework.authtoken.models import Token
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.generics import CreateAPIView
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.viewsets import ModelViewSet
from django.views import View
from rest_framework import viewsets
from .models import CustomUser, FinancialDetails, ForecastedFinancialDetails,FinancialData
from .serializers import (FinancialDetailsSerializer,
                          ForecastedFinancialDetailsSerializer, UserSerializer,FinancialDataSerializer,FinancialDataQuerySerializer,CustomUserSerializer)


#Register API
class UserRegistrationView(CreateAPIView):
    queryset = CustomUser.objects.all()
    serializer_class = UserSerializer

    def get(self, request, *args, **kwargs):
        return render(request, 'register.html')

#Login API
@method_decorator(csrf_exempt, name='dispatch')
class UserLoginView(APIView):
    template_name = 'login.html'

    def get(self, request, *args, **kwargs):
        return render(request, self.template_name, {})

    def post(self, request, *args, **kwargs):
        username = request.data.get('username')
        password = request.data.get('password')
        print(f"Received username: {username}, password: {password}")

        user = authenticate(request, username=username, password=password)
        print(f"Authenticated user: {user}")

        if user is not None:
            login(request, user)

            # Delete existing tokens !!!
            Token.objects.filter(user=user).delete()

            # Create a new token !
            token = Token.objects.create(user=user)

            return redirect(reverse('home'))
        else:
            return redirect('login')




#logout
# @method_decorator(csrf_exempt, name='dispatch')
# class UserLogoutView(APIView):
#     def post(self, request, *args, **kwargs):
#         logout(request)
#         return Response({'message': 'User logged out successfully'}, status=status.HTTP_200_OK)


#Mini Questionaries API's
        
class FinancialDetailsView(TemplateView, generics.CreateAPIView):
    template_name = 'form3.html'
    serializer_class = FinancialDetailsSerializer
    # permission_classes = [IsAuthenticated]

    def create(self, request, *args, **kwargs):
        # Ensure only one FinancialDetails record per user
        if FinancialDetails.objects.filter(user=request.user).exists():
            return Response({'error': 'FinancialDetails record already exists for this user.'}, status=400)
        return super().create(request, *args, **kwargs)

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)






class ForecastedFinancialDetailsView(TemplateView,generics.CreateAPIView):
    template_name='form3.html'
    serializer_class = ForecastedFinancialDetailsSerializer
    permission_classes = [IsAuthenticated]

    def create(self, request, *args, **kwargs):
        # Ensure only one ForecastedFinancialDetails record per user
        if ForecastedFinancialDetails.objects.filter(user=request.user).exists():
            return Response({'error': 'ForecastedFinancialDetails record already exists for this user.'}, status=400)
        return super().create(request, *args, **kwargs)

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)




#Home
@login_required
def home(request):
    context = {'message': 'Welcome to the Home Page!'}
    return render(request, 'carousal.html', context)



#FinancialData

# class FinancialDataView(ModelViewSet):
#     serializer_class=FinancialDataSerializer
#     queryset=FinancialData.objects.all()


    # @action(detail=False,methods=["get"])
    # def download(self,request):
    #     queryset=self.get_queryset()
    #     serializer=FinancialDataSerializer(queryset,many=True)

    #     return Response(serializer.data)

class CalculateMetricsView(APIView):
    permission_classes = [IsAuthenticated]
    
    # @method_decorator(csrf_exempt)
    # def get(self,request):

    #     return render(request,'template')
    
    @method_decorator(csrf_exempt)
    def post(self, request):
        metric_requested = request.data.get('metric')
        
        print(f"Received metric_requested: {metric_requested}")

        if metric_requested not in ['Sales', 'PAT', 'Shares Issued', 'Book Value', 'EBITDA']:
            return Response({'error': 'Invalid metric requested.'}, status=status.HTTP_400_BAD_REQUEST)

        companies = CustomUser.objects.all().annotate(
            cmp_rs=F('financialdetails__pat') * 15,  # PE ratio 15
            mar_cap_rs_cr=(F('financialdetails__pat') * F('financialdetails__book_value')) / 10000000 
        )

        if metric_requested == 'Sales':
            sorted_companies = companies.order_by('-financialdetails__sales')[:5]
        elif metric_requested == 'PAT':
            sorted_companies = companies.order_by('-cmp_rs')[:5]
        elif metric_requested == 'Shares Issued':
            sorted_companies = companies.order_by('-financialdetails__shares_issued')[:5]
        elif metric_requested == 'Book Value':
            sorted_companies = companies.order_by('-financialdetails__book_value')[:5]
        else:  
            sorted_companies = companies.order_by('-financialdetails__ebitda')[:5]

        if not sorted_companies:
            return Response({'error': 'No companies found for the specified metric category.'}, status=status.HTTP_404_NOT_FOUND)

        response_data = []
        for company in sorted_companies:
            financial_details = company.financialdetails  # Access FinancialDetails 

            selected_metric_value = getattr(financial_details, metric_requested.lower().replace(" ", "_")) / 10000000  # Convert to Crores

            response_data.append({
                'Name of the company': company.startup_name,
                'CMP Rs.': company.cmp_rs,
                'Mar Cap Rs.Cr.': company.mar_cap_rs_cr,
                f'{metric_requested} in Cr.': selected_metric_value
            })

        return Response(response_data, status=status.HTTP_200_OK)    
    
    
    
    
#Download API:-[input,industry data,ratios derived,output]            
class DownloadTopCompaniesExcelView(View):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user
        top_companies = CustomUser.objects.all().annotate(
            cmp_rs=F('financialdetails__pat') * 15,
            mar_cap_rs_cr=(F('financialdetails__pat') * F('financialdetails__book_value')) / 10000000
        ).order_by('-financialdetails__sales')[:5]

        if not top_companies:
            return HttpResponse('No companies found for the specified metric category.', status=404)
        #Data Frames
        top_companies_data = {
            'Name of the company': [],
            'CMP Rs.': [],
            'Mar Cap Rs.Cr.': [],
            'Metric': [],
        }

        current_user_data = {
            'Metric': ['Sales', 'EBITDA', 'PAT', 'Shares Issued', 'Book Value'],
            'Metric Value': [user.financialdetails.sales,
                             user.financialdetails.ebitda,
                             user.financialdetails.pat,
                             user.financialdetails.shares_issued,
                             user.financialdetails.book_value],
        }

        for metric in ['Sales', 'EBITDA', 'PAT', 'Shares Issued', 'Book Value']:
            for company in top_companies:
                financial_details = company.financialdetails

                top_companies_data['Name of the company'].append(company.startup_name)
                top_companies_data['CMP Rs.'].append(company.cmp_rs)
                top_companies_data['Mar Cap Rs.Cr.'].append(company.mar_cap_rs_cr)
                top_companies_data['Metric'].append(f'{metric}')

            # empty rows
            if metric == 'Sales':
                for _ in range(2):
                    top_companies_data['Name of the company'].append('')
                    top_companies_data['CMP Rs.'].append('')
                    top_companies_data['Mar Cap Rs.Cr.'].append('')
                    top_companies_data['Metric'].append('')
            elif metric == 'EBITDA':
                for _ in range(2):
                    top_companies_data['Name of the company'].append('')
                    top_companies_data['CMP Rs.'].append('')
                    top_companies_data['Mar Cap Rs.Cr.'].append('')
                    top_companies_data['Metric'].append('')
            elif metric == 'PAT':
                for _ in range(2):
                    top_companies_data['Name of the company'].append('')
                    top_companies_data['CMP Rs.'].append('')
                    top_companies_data['Mar Cap Rs.Cr.'].append('')
                    top_companies_data['Metric'].append('')
            elif metric == 'Shares Issued':
                for _ in range(2):
                    top_companies_data['Name of the company'].append('')
                    top_companies_data['CMP Rs.'].append('')
                    top_companies_data['Mar Cap Rs.Cr.'].append('')
                    top_companies_data['Metric'].append('')
            else:
                pass
                    
                

        # Additional frame 
        industry_frame_data = {
            'Metric': ['Industry'],
            'Metric Value': [user.financialdetails.industry],
        }

        ratios_derived_data = {
            'Metric': ['Average PE', 'Average PS', 'Average P to BV', 'Average P to CF from OA', 'Average P to EBITDA'],
            'Metric Value': [55.12596748,1.254577347,4.172895978,31.55355969,11.71104351] 
        }

        df_top_companies = pd.DataFrame(top_companies_data)
        df_current_user = pd.DataFrame(current_user_data)
        df_industry_frame = pd.DataFrame(industry_frame_data)
        df_ratios_derived_frame = pd.DataFrame(ratios_derived_data)

        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df_top_companies.to_excel(writer, sheet_name='Top Companies Metrics', index=False)
            df_current_user.to_excel(writer, sheet_name='Input Metrics', index=False)
            df_industry_frame.to_excel(writer, sheet_name=f'Industry - {user.financialdetails.industry}', index=False)
            df_ratios_derived_frame.to_excel(writer, sheet_name='Ratios Derived', index=False)

        output.seek(0)

        response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=top_companies_metrics.xlsx'

        return response
    
#top 5 companies data



class FinancialDataView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, *args, **kwargs):
        try:
            query = request.data.get('query', '')
            conditions = query.split('and')
            dynamic_query = Q()
            operators_mapping = {'>': 'gt', '<': 'lt', '>=': 'gte', '<=': 'lte', '=': 'exact'}
            
            for condition in conditions:
                condition_parts = condition.strip().split()
                if len(condition_parts) != 3:
                    raise ValueError("Invalid condition format. Please use 'field operator value'.")
                field, operator, value = condition_parts
                django_operator = operators_mapping.get(operator)
                if not django_operator:
                    raise ValueError(f"Invalid operator: {operator}")
                dynamic_query &= Q(**{f"{field}__{django_operator}": float(value)})
            
            # Sample Queries
            # 1. Companies with Market Capitalization > 500 and Price to Earning Ratio < 15
            # dynamic_query &= Q(mar_cap_rs__gt=500, pe__lt=15)

            # 2. Companies in the 'Technology' industry with Profit Growth > 5
            # dynamic_query &= Q(industry='Technology', profit_growth_percentage__gt=5)

            results = FinancialData.objects.filter(dynamic_query).order_by('cmp_rs')[:15]
            
            serialized_results = [
                {
                    'name': entry.name,
                    'cmp_rs': entry.cmp_rs,
                    'pe': entry.pe,
                    'mar_cap_rs': entry.mar_cap_rs,
                    'num_of_shares': entry.num_of_shares,
                    'np_qtr_rs': entry.np_qtr_rs,
                    'sales_qtr_rs': entry.sales_qtr_rs,
                    'roce_percentage': entry.roce_percentage,
                    'sales_rs': entry.sales_rs,
                    'sales_per_share': entry.sales_per_share,
                    'cmp_over_sales': entry.cmp_over_sales,
                    'cmp_over_bv': entry.cmp_over_bv,
                    'bv_in_cr': entry.bv_in_cr,
                    'ind_pe': entry.ind_pe,
                    'profit_growth_percentage': entry.profit_growth_percentage,
                    'ev_rs': entry.ev_rs,
                    'pat_12m_rs': entry.pat_12m_rs,
                    'debt_eq': entry.debt_eq,
                    'cmp_over_ocf': entry.cmp_over_ocf,
                    'ev_over_ebitda': entry.ev_over_ebitda,
                    'industry': entry.industry,
                } for entry in results
            ]
            
            return Response(serialized_results, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


class FinancialDataScreenerView(APIView):
    permission_classes = [IsAuthenticated]

    template_name = 'Screener.html'

    def get(self, request, *args, **kwargs):
        context = {}
        return render(request, self.template_name, context)

class CustomUserView(RetrieveUpdateAPIView):
    queryset = CustomUser.objects.all()
    serializer_class = CustomUserSerializer
    permission_classes = [IsAuthenticated]
    template_name = '#'  

    def get_object(self):
        return self.request.user

    def get(self, request, *args, **kwargs):
        user_instance = self.get_object()

        additional_context = {
            'user_data': self.get_serializer(user_instance).data,
            'custom_data': 'Some additional data you want to pass',
        }

        return render(request, self.template_name, context=additional_context)    
    

# from django.db.models import Q
# #Getting all data
# class FinancialDataView(APIView):
#     def post(self, request, *args, **kwargs):
#         try:
#             query = request.data.get('query', '')
            
#             # Split the query into field, operator, and value
#             query_parts = query.split()

#             # Check if the split result has exactly three parts
#             if len(query_parts) != 3:
#                 raise ValueError("Invalid query format. Please use 'field operator value'.")

#             field, operator, value = query_parts

#             # Define a dictionary mapping operators to their Django ORM equivalents
#             operators_mapping = {'>': 'gt', '<': 'lt', '>=': 'gte', '<=': 'lte', '=': 'exact'}
            
#             # Check if the operator is valid
#             django_operator = operators_mapping.get(operator)
#             if not django_operator:
#                 raise ValueError(f"Invalid operator: {operator}")

#             # Construct the dynamic query using Django ORM
#             filter_kwargs = {f"{field}__{django_operator}": float(value)}
#             results = FinancialData.objects.filter(**filter_kwargs)

#             # You might want to serialize the results before sending them in the response.
#             # For simplicity, I'm just converting them to a list for demonstration purposes.
#             serialized_results = [{'name': entry.name, 'cmp_rs': entry.cmp_rs, 'pe': entry.pe} for entry in results]

#             return Response(serialized_results, status=status.HTTP_200_OK)

#         except Exception as e:
#             return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        




# class FinancialDataView(APIView):
#     def post(self, request, *args, **kwargs):
#         try:
#             input_data = request.data
#             query_conditions = input_data.get("query", [])
#             limit = input_data.get("limit", None)

#             if not isinstance(query_conditions, list):
#                 raise ValueError("Invalid query format. Please provide a list of conditions.")

#             # Build queryset based on conditions
#             queryset = FinancialData.objects.all()
#             for condition in query_conditions:
#                 if not isinstance(condition, dict):
#                     raise ValueError("Invalid condition format. Please provide a dictionary.")

#                 field = condition.get("field")
#                 operator = condition.get("operator")
#                 value = condition.get("value")

#                 if field is None or operator is None or value is None:
#                     raise ValueError("Invalid condition format. Please include 'field', 'operator', and 'value'.")

#                 # Process the condition and filter the queryset
#                 condition_filter = f"{field}__{operator}"
#                 queryset = queryset.filter(**{condition_filter: value})

#             # Apply limit to queryset if needed
#             if limit is not None:
#                 queryset = queryset[:limit]

#             # Serialize the data
#             serializer = FinancialDataSerializer(queryset, many=True)

#             # Return the response
#             return Response({"success": True, "data": serializer.data})
#         except Exception as e:
#             return Response({"error": str(e)})
    

# class FinancialDataView(APIView):
#     def post(self, request):
#         query_string = request.GET.get('query', '')

#         # Split the query into individual conditions
#         conditions = [condition.strip() for condition in query_string.split('and')]

#         # Build Q objects for each condition
#         q_objects = []
#         for condition in conditions:
#             parts = condition.split()
            
#             # Check if the condition has the expected format
#             if len(parts) == 3:
#                 field, operator, value = parts
#                 q_objects.append(Q(**{f"{field}__{operator}": value}))
#             else:
#                 return Response({"error": "Invalid condition format"})

#         # Combine Q objects with AND
#         query_filter = Q()
#         for q_object in q_objects:
#             query_filter &= q_object

#         # Retrieve data based on the constructed filter
#         queryset = FinancialData.objects.filter(query_filter)[:5]  # Limit to top 5
#         serializer = FinancialDataSerializer(queryset, many=True)

#         return Response(serializer.data)


















#Uploading the excel data
# class ExcelImportView(APIView):
#     template_name = 'Templates/excel_import.html'
#     parser_classes = (MultiPartParser,)

#     def get(self, request, *args, **kwargs):
#         form = ExcelUploadForm()
#         return render(request, 'excel_import.html', {'form': form})

#     def post(self, request, *args, **kwargs):
#         form = ExcelUploadForm(request.POST, request.FILES)

#         if form.is_valid():
#             try:
#                 excel_file = request.FILES['excel_file']
#                 self.handle_excel(excel_file)
#                 return Response({'message': 'File uploaded and processed successfully.'}, status=status.HTTP_201_CREATED)
#             except Exception as e:
#                 print(f"Error processing Excel file: {e}")
#                 return Response({'error': 'Failed to process the Excel file.'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
#         else:
#             return render(request, self.template_name, {'form': form})

#     def handle_excel(self, excel_file):
#         workbook = openpyxl.load_workbook(excel_file, data_only=True)
#         sheet = workbook.active

#         for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
#             try:
#                 with transaction.atomic():
#                     FinancialData.objects.create(
#                         name=row[0],
#                         cmp_rs=row[1],
#                         pe=row[2],
#                         mar_cap_rs=row[3],
#                         num_of_shares=row[4],
#                         np_qtr_rs=row[5],
#                         sales_qtr_rs=row[6],
#                         roce_percentage=row[7],
#                         sales_rs=row[8],
#                         sales_per_share=row[9],
#                         cmp_over_sales=row[10],
#                         cmp_over_bv=row[11],
#                         bv_in_cr=row[12],
#                         ind_pe=row[13],
#                         profit_growth_percentage=row[14],
#                         ev_rs=row[15],
#                         pat_12m_rs=row[16],
#                         debt_eq=row[17],
#                         cmp_over_ocf=row[18],
#                         ev_over_ebitda=row[19],
#                         industry=row[20]
#                     )
#                     print("Transaction committed successfully")

#             except Exception as e:
#                 print(f"Error processing row {row_num}: {e}")

# class ExcelImportView(APIView):
#     template_name = 'Templates/excel_import.html'
#     parser_classes = (MultiPartParser,)

#     def get(self, request, *args, **kwargs):
#         form = ExcelUploadForm()
#         return render(request, 'excel_import.html', {'form': form})

#     def post(self, request, *args, **kwargs):
#         form = ExcelUploadForm(request.POST, request.FILES)

#         if form.is_valid():
#             try:
#                 excel_file = request.FILES['excel_file']
#                 self.handle_excel(excel_file)
#                 return Response({'message': 'File uploaded and processed successfully.'}, status=status.HTTP_201_CREATED)
#             except Exception as e:
#                 print(f"Error processing Excel file: {e}")
#                 return Response({'error': 'Failed to process the Excel file.'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
#         else:
#             return render(request, self.template_name, {'form': form})

#     def handle_excel(self, excel_file):
#         workbook = openpyxl.load_workbook(excel_file, data_only=True)
#         sheet = workbook.active

#         for row in sheet.iter_rows(min_row=2, values_only=True):
#             try:
#                 with transaction.atomic():
#                     FormulaValue.objects.create(
#                         revised_tag=row[0],
#                         symbol=row[1],
#                         audited=row[2] == 'U',  
#                         standalone=row[3] == 'S',  
#                         value=row[4],
#                         period=row[5]
#                     )
#                     print("Transaction committed successfully")
#             except Exception as e:
#                 print(f"Error processing row {row}: {e}")
